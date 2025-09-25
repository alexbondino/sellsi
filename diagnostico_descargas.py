#!/usr/bin/env python3
"""
Script de Diagnóstico de Descargas
Analiza la carpeta de descargas para verificar si se descargaron todos los archivos esperados
"""

import json
from pathlib import Path
from openpyxl import load_workbook


def diagnosticar_descargas(directorio_descargas):
    """
    Analiza la carpeta de descargas y verifica:
    1. Cuántos archivos se descargaron
    2. Si existe ContratoSwapTasas.xlsx
    3. Si los archivos tienen hoja 'Swap' con celda C8
    4. Qué valores están en C8
    """

    directorio = Path(directorio_descargas)
    if not directorio.exists():
        print(f"❌ ERROR: Directorio {directorio} no existe")
        return False

    archivos_excel = list(directorio.glob("*.xlsx"))
    print(f"📁 Directorio: {directorio}")
    print(f"📊 Total archivos Excel encontrados: {len(archivos_excel)}")

    # Verificar ContratoSwapTasas.xlsx
    contrato_swap = directorio / "ContratoSwapTasas.xlsx"
    if contrato_swap.exists():
        print("✅ ContratoSwapTasas.xlsx: ENCONTRADO")
    else:
        print("❌ ContratoSwapTasas.xlsx: NO ENCONTRADO")

    # Analizar archivos de contratos
    print("\n🔍 Análisis de archivos de contratos:")
    bancos_encontrados = {}
    archivos_sin_swap = []
    archivos_sin_c8 = []

    for archivo in archivos_excel:
        if archivo.name.lower() == "contratoswaptasas.xlsx":
            continue

        print(f"\n📄 Analizando: {archivo.name}")

        try:
            wb = load_workbook(archivo, read_only=True)

            # Verificar hoja 'Swap'
            if "Swap" not in wb.sheetnames:
                print("  ❌ No tiene hoja 'Swap'")
                archivos_sin_swap.append(archivo.name)
                continue

            ws = wb["Swap"]

            # Leer celda C8
            valor_c8 = ws["C8"].value
            if valor_c8:
                print(f"  ✅ Celda C8: '{valor_c8}'")
                bancos_encontrados[valor_c8] = archivo.name
            else:
                print("  ❌ Celda C8 está vacía")
                archivos_sin_c8.append(archivo.name)

            wb.close()

        except Exception as e:
            print(f"  ❌ Error leyendo archivo: {e}")

    # Verificar contra el mapeo del macro
    bancos_esperados = {
        "Scotiabank N°3",
        "BCI N°1",
        "Scotiabank N°1",
        "Santander N°2",
        "Scotiabank N°2",
        "BCI N°2",
        "Scotiabank N°4",
        "Santander N°3",
        "Santander N°4",
        "BCI N°3",
        "Scotiabank N°5",
        "Santander N°1",
    }

    print("\n📋 RESUMEN DE BANCOS:")
    bancos_encontrados_set = set(bancos_encontrados.keys())
    bancos_faltantes = bancos_esperados - bancos_encontrados_set
    bancos_extra = bancos_encontrados_set - bancos_esperados

    print(
        f"✅ Bancos encontrados ({len(bancos_encontrados_set)}): {sorted(bancos_encontrados_set)}"
    )

    if bancos_faltantes:
        print(f"❌ Bancos FALTANTES ({len(bancos_faltantes)}): {sorted(bancos_faltantes)}")
    else:
        print("✅ Todos los bancos esperados están presentes")

    if bancos_extra:
        print(f"⚠️ Bancos EXTRA ({len(bancos_extra)}): {sorted(bancos_extra)}")

    # Verificar reporte de descarga si existe
    reporte_path = directorio / "mtm_download_report.json"
    if reporte_path.exists():
        print("\n📊 REPORTE DE DESCARGAS:")
        try:
            with open(reporte_path, "r", encoding="utf-8") as f:
                reporte = json.load(f)

            summary = reporte.get("summary", {})
            print(f"✅ Exitosas: {summary.get('success', 0)}")
            print(f"❌ Fallidas: {summary.get('failed', 0)}")
            print(f"⚠️ Inválidas: {summary.get('invalid', 0)}")

            # Mostrar fallos específicos
            sessions = reporte.get("sessions", [])
            fallos = [s for s in sessions if s["status"] == "failed"]
            if fallos:
                print("\n❌ FALLOS DETECTADOS:")
                for fallo in fallos:
                    contract = fallo.get("metadata", {}).get("contract_name", "unknown")
                    reason = fallo.get("metadata", {}).get("failure_reason", "unknown")
                    print(f"  - {contract}: {reason}")

        except Exception as e:
            print(f"❌ Error leyendo reporte: {e}")
    else:
        print("\n⚠️ No se encontró reporte de descargas (mtm_download_report.json)")

    print(f"\n{'='*60}")
    print("DIAGNÓSTICO FINAL:")

    total_esperado = 13  # 1 ContratoSwapTasas + 12 contratos
    total_encontrado = len(archivos_excel)

    if (
        total_encontrado == total_esperado
        and not bancos_faltantes
        and not archivos_sin_swap
        and not archivos_sin_c8
    ):
        print("✅ PERFECTO: Todos los archivos están presentes y correctos")
        return True
    else:
        print("❌ PROBLEMAS DETECTADOS:")
        if total_encontrado != total_esperado:
            print(f"  - Archivos: esperados {total_esperado}, encontrados {total_encontrado}")
        if bancos_faltantes:
            print(f"  - Faltan bancos: {len(bancos_faltantes)}")
        if archivos_sin_swap:
            print(f"  - Archivos sin hoja Swap: {len(archivos_sin_swap)}")
        if archivos_sin_c8:
            print(f"  - Archivos sin valor en C8: {len(archivos_sin_c8)}")
        return False


if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1:
        directorio = sys.argv[1]
    else:
        directorio = input("Ingresa la ruta del directorio de descargas: ").strip()

    diagnosticar_descargas(directorio)
