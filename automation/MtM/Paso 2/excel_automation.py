"""
Excel Automation Module - Paso 2 MtM
Autor: GitHub Copilot
Fecha: 29 de Septiembre, 2025

Módulo especializado para automatización de Excel en el Paso 2 del proceso MtM.
Maneja la escritura automática en celda I14 y ejecución de macros VBA.
"""

import os
import time
from typing import Optional, Callable


class ExcelAutomator:
    """
    Automatizador de Excel para el proceso MtM Paso 2
    
    Funcionalidades:
    - Escritura automática en celda I14 (carpeta de descarga)
    - Ejecución automática de macros VBA
    - Manejo de errores y fallbacks
    - Logging integrado
    """
    
    def __init__(self, log_callback: Optional[Callable] = None):
        """
        Inicializa el automatizador de Excel
        
        Args:
            log_callback: Función de logging para reportar progreso
        """
        self.log = log_callback if log_callback else self._default_log
        
    def _default_log(self, message: str, level: str = "INFO"):
        """Log por defecto si no se proporciona callback"""
        print(f"[{level}] {message}")
    
    def escribir_carpeta_en_celda(self, excel_file_path: str, carpeta_descarga: str) -> bool:
        """
        Escribe automáticamente la carpeta de descarga en la celda I14
        
        Args:
            excel_file_path: Ruta completa del archivo Excel (.xlsm)
            carpeta_descarga: Ruta de la carpeta de descarga
            
        Returns:
            bool: True si la escritura fue exitosa, False en caso contrario
        """
        try:
            self.log("🔧 [DEBUG] Iniciando escritura automática en Excel", level="DEBUG")
            self.log(f"🔧 [DEBUG] Archivo: {excel_file_path}", level="DEBUG")
            self.log(f"🔧 [DEBUG] Carpeta: {carpeta_descarga}", level="DEBUG")
            
            # Verificar que el archivo existe
            if not os.path.exists(excel_file_path):
                self.log(f"❌ Archivo Excel no existe: {excel_file_path}", level="ERROR")
                return False
            
            # Intentar usar openpyxl (método preferido)
            try:
                from openpyxl import load_workbook
                
                self.log("📝 Escribiendo carpeta de descarga en Excel...", level="INFO")
                
                # Cargar el workbook preservando VBA
                wb = load_workbook(excel_file_path, keep_vba=True)
                self.log("🔧 [DEBUG] Workbook cargado exitosamente", level="DEBUG")
                
                # Buscar hoja "Contratos" (case-insensitive)
                hoja_contratos = None
                for hoja_name in wb.sheetnames:
                    if hoja_name.lower() == "contratos":
                        hoja_contratos = hoja_name
                        break
                
                if not hoja_contratos:
                    self.log("⚠️ Hoja 'Contratos' no encontrada en Excel", level="ADVERTENCIA")
                    self.log(f"📋 Hojas disponibles: {', '.join(wb.sheetnames)}", level="INFO")
                    wb.close()
                    return False
                
                # Seleccionar hoja y escribir en I14
                ws = wb[hoja_contratos]
                valor_anterior = ws["I14"].value
                
                ws["I14"] = carpeta_descarga
                self.log(f"🔧 [DEBUG] I14: '{valor_anterior}' → '{carpeta_descarga}'", level="DEBUG")
                
                # Guardar cambios
                wb.save(excel_file_path)
                wb.close()
                
                self.log(f"✅ Carpeta configurada en I14: {carpeta_descarga}", level="EXITO")
                return True
                
            except ImportError:
                self.log("⚠️ openpyxl no disponible", level="ADVERTENCIA")
                self.log("💡 Instala openpyxl: pip install openpyxl", level="INFO")
                return False
            except Exception as e:
                self.log(f"❌ Error con openpyxl: {type(e).__name__}: {e}", level="ERROR")
                return False
                
        except Exception as e:
            self.log(f"❌ Error general escribiendo en Excel: {e}", level="ERROR")
            return False
    
    def ejecutar_macro_vba(self, excel_file_path: str, modulo_vba: str, macro_name: str) -> bool:
        """
        Ejecuta un macro VBA específico usando COM automation
        
        Args:
            excel_file_path: Ruta completa del archivo Excel (.xlsm)
            modulo_vba: Nombre del módulo VBA (ej: "LimpiarTablas")
            macro_name: Nombre del macro (ej: "LimpiarTablasContratos")
            
        Returns:
            bool: True si la ejecución fue exitosa, False en caso contrario
        """
        try:
            self.log(f"⚡ Preparando ejecución de macro: {modulo_vba}.{macro_name}", level="INFO")
            
            try:
                import win32com.client
                
                self.log("📋 Conectando con Excel via COM...", level="INFO")
                
                # Conectar con Excel
                excel_app = win32com.client.Dispatch("Excel.Application")
                excel_app.Visible = True  # Mostrar Excel para feedback visual
                excel_app.DisplayAlerts = False  # Evitar diálogos durante ejecución
                
                # Abrir archivo Excel
                self.log(f"📂 Abriendo archivo: {os.path.basename(excel_file_path)}", level="INFO")
                workbook = excel_app.Workbooks.Open(excel_file_path)
                
                # Esperar carga completa
                time.sleep(2)
                
                # Ejecutar macro
                macro_completo = f"{modulo_vba}.{macro_name}"
                self.log(f"🚀 Ejecutando macro: {macro_completo}", level="INFO")
                excel_app.Run(macro_completo)
                
                # Guardar cambios automáticamente
                self.log("💾 Guardando cambios automáticamente...", level="INFO")
                workbook.Save()
                
                # Restaurar alertas pero NO cerrar Excel
                excel_app.DisplayAlerts = True
                self.log("✅ Macro ejecutado exitosamente", level="EXITO")
                self.log("📋 Excel permanece abierto para trabajo manual", level="INFO")
                
                return True
                
            except ImportError:
                self.log("⚠️ win32com no disponible", level="ADVERTENCIA")
                self.log("💡 Instala pywin32: pip install pywin32", level="INFO")
                return False
            except Exception as e:
                self.log(f"❌ Error ejecutando macro: {type(e).__name__}: {e}", level="ERROR")
                self.log(f"🔧 [DEBUG] Detalles: {str(e)}", level="DEBUG")
                return False
                
        except Exception as e:
            self.log(f"❌ Error general en ejecución de macro: {e}", level="ERROR")
            return False
    
    def proceso_completo_paso2(self, excel_file_path: str, carpeta_descarga: str, 
                              modulo_vba: str = "LimpiarTablas", 
                              macro_name: str = "LimpiarTablasContratos") -> bool:
        """
        Ejecuta el proceso completo del Paso 2: Escritura + Macros en secuencia
        
        Args:
            excel_file_path: Ruta del archivo Excel MtM
            carpeta_descarga: Carpeta de descarga a configurar
            modulo_vba: Módulo VBA que contiene el primer macro
            macro_name: Nombre del primer macro a ejecutar
            
        Returns:
            bool: True si todo el proceso fue exitoso
        """
        self.log("🚀 INICIANDO PROCESO COMPLETO PASO 2", level="INFO")
        
        # Paso 1: Escribir carpeta en I14 (archivo cerrado)
        self.log("📝 FASE 1: Configurando carpeta en celda I14...", level="INFO")
        if not self.escribir_carpeta_en_celda(excel_file_path, carpeta_descarga):
            self.log("❌ Error en Fase 1 - Abortando proceso", level="ERROR")
            return False
        
        # Paso 2: Ejecutar macro de limpieza (archivo abierto)
        self.log("⚡ FASE 2: Ejecutando macro de limpieza automática...", level="INFO")
        if not self.ejecutar_macro_vba(excel_file_path, modulo_vba, macro_name):
            self.log("❌ Error en Fase 2 - Proceso incompleto", level="ERROR")
            # Intentar abrir Excel manualmente como fallback
            self._fallback_abrir_excel(excel_file_path)
            return False
        
        # Paso 3: Ejecutar macro de seleccionar directorio
        self.log("📂 FASE 3: Ejecutando macro de selección de directorio...", level="INFO")
        if not self.ejecutar_macro_vba(excel_file_path, "CopiarTablas", "SeleccionarDirectorioGuardarEnCeldas"):
            self.log("❌ Error en Fase 3 - Proceso incompleto", level="ERROR")
            self.log("⚠️ Macro SeleccionarDirectorioGuardarEnCeldas falló", level="ADVERTENCIA")
            # Excel ya está abierto, el usuario puede continuar manualmente
            return False
        
        # Proceso completado exitosamente
        self.log("🎉 ¡PROCESO PASO 2 COMPLETADO EXITOSAMENTE!", level="EXITO")
        self.log("✅ ✓ Carpeta configurada automáticamente", level="EXITO")
        self.log("✅ ✓ Tablas limpiadas automáticamente", level="EXITO")
        self.log("✅ ✓ Directorio seleccionado y guardado", level="EXITO")
        self.log("✅ ✓ Excel listo para trabajo manual", level="EXITO")
        
        return True
    
    def _fallback_abrir_excel(self, excel_file_path: str):
        """Fallback: Abrir Excel manualmente si falla la automatización"""
        try:
            self.log("🔄 Fallback: Abriendo Excel manualmente...", level="INFO")
            os.startfile(excel_file_path)
            self.log("📂 Excel abierto - configura manualmente si es necesario", level="INFO")
        except Exception as e:
            self.log(f"❌ Fallback falló: {e}", level="ERROR")


def ejecutar_automatizacion_excel(excel_file_path: str, carpeta_descarga: str, 
                                 log_callback: Optional[Callable] = None) -> bool:
    """
    Función principal para ejecutar automatización de Excel en Paso 2
    
    Args:
        excel_file_path: Ruta completa del archivo Excel MtM
        carpeta_descarga: Carpeta de descarga a configurar
        log_callback: Función de logging (opcional)
        
    Returns:
        bool: True si el proceso fue exitoso
    """
    automator = ExcelAutomator(log_callback)
    return automator.proceso_completo_paso2(excel_file_path, carpeta_descarga)


if __name__ == "__main__":
    # Test básico del módulo
    print("🧪 Módulo Excel Automation - Paso 2 MtM")
    print("✅ Importación exitosa")