"""
Test de generación de Excel con NUEVA estructura spot_fijo
Sin ejecutar el bot completo, solo genera el Excel con datos de prueba
"""

from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Datos de prueba con NUEVA ESTRUCTURA (spot_fijo en iteración 1)
resultados_matriz = {
    "30/10/2025": {
        "fecha": datetime(2025, 10, 30),
        # ITERACIÓN 1: spot_fijo capturado
        "UF/USD": {"spot_fijo": "0,0254", "pts_fwd": "0,0001"},
        "USD/CLP": {"spot_fijo": "920,50", "pts_fwd": "12,30"},
        "EUR/CLP": {"spot_fijo": "1050,20", "pts_fwd": "15,40"}
    },
    "28/11/2025": {
        "fecha": datetime(2025, 11, 28),
        # ITERACIÓN 2+: Solo pts_fwd
        "UF/USD": {"spot_fijo": None, "pts_fwd": "0,0002"},
        "USD/CLP": {"spot_fijo": None, "pts_fwd": "13,10"},
        "EUR/CLP": {"spot_fijo": None, "pts_fwd": "16,20"}
    },
    "30/12/2025": {
        "fecha": datetime(2025, 12, 30),
        # ITERACIÓN 3+: Solo pts_fwd
        "UF/USD": {"spot_fijo": None, "pts_fwd": "0,0003"},
        "USD/CLP": {"spot_fijo": None, "pts_fwd": "14,50"},
        "EUR/CLP": {"spot_fijo": None, "pts_fwd": "17,80"}
    }
}

paridades = ["USD/CLP", "EUR/CLP", "UF/USD"]
paridades_normales = ["USD/CLP", "EUR/CLP"]
tiene_uf_usd = True

# Crear workbook
wb = Workbook()
ws = wb.active
ws.title = "Forward Results TEST"

print("✅ Generando Excel con NUEVA estructura spot_fijo...")

# ===== ROW 1: HEADERS MERGED =====
# Columnas B-E: "UF y IPC"
cell = ws.cell(row=1, column=2, value="UF y IPC")
cell.font = Font(bold=True, size=12, color="FFFFFF")
cell.alignment = Alignment(horizontal='center', vertical='center')
cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=5)

# Columnas F-H: "USD"
cell = ws.cell(row=1, column=6, value="USD")
cell.font = Font(bold=True, size=12, color="FFFFFF")
cell.alignment = Alignment(horizontal='center', vertical='center')
cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=8)

# Columnas I-K: "EUR"
cell = ws.cell(row=1, column=9, value="EUR")
cell.font = Font(bold=True, size=12, color="FFFFFF")
cell.alignment = Alignment(horizontal='center', vertical='center')
cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=11)

# ===== ROW 2: SUB-HEADERS =====
header_style = {
    'font': Font(bold=True, size=10),
    'fill': PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
    'alignment': Alignment(horizontal='center', vertical='center')
}

headers_row2 = [
    (1, "Fecha"),
    (2, "Spot"),
    (3, "Pts. Forward"),
    (4, "Precio Forward"),
    (5, "IPC"),
    (6, "Spot"),
    (7, "Pts. Forward"),
    (8, "Precio Forward"),
    (9, "Spot"),
    (10, "Pts. Forward"),
    (11, "Precio Forward")
]

for col, text in headers_row2:
    cell = ws.cell(row=2, column=col, value=text)
    for k, v in header_style.items():
        setattr(cell, k, v)

# ===== EXTRAER SPOTS FIJOS =====
fechas_ordenadas = sorted(resultados_matriz.keys(), key=lambda x: resultados_matriz[x]["fecha"])

spot_fijo_uf = None
spot_fijo_usd = None
spot_fijo_eur = None

for fecha_str in fechas_ordenadas:
    data_row = resultados_matriz[fecha_str]
    
    if "UF/USD" in data_row and data_row["UF/USD"].get("spot_fijo") and spot_fijo_uf is None:
        spot_fijo_uf = data_row["UF/USD"]["spot_fijo"]
    
    if "USD/CLP" in data_row and data_row["USD/CLP"].get("spot_fijo") and spot_fijo_usd is None:
        spot_fijo_usd = data_row["USD/CLP"]["spot_fijo"]
    
    if "EUR/CLP" in data_row and data_row["EUR/CLP"].get("spot_fijo") and spot_fijo_eur is None:
        spot_fijo_eur = data_row["EUR/CLP"]["spot_fijo"]

print(f"📌 Spots fijos extraídos: UF={spot_fijo_uf}, USD={spot_fijo_usd}, EUR={spot_fijo_eur}")

# ===== ROW 3+: FILAS DE DATOS =====
# Convertir spots fijos a números ANTES del loop
spot_fijo_uf_num = None
spot_fijo_usd_num = None
spot_fijo_eur_num = None

if spot_fijo_uf:
    try:
        spot_fijo_uf_num = float(spot_fijo_uf.replace(',', '.'))
    except (ValueError, AttributeError):
        pass

if spot_fijo_usd:
    try:
        spot_fijo_usd_num = float(spot_fijo_usd.replace('.', '').replace(',', '.'))
    except (ValueError, AttributeError):
        pass

if spot_fijo_eur:
    try:
        spot_fijo_eur_num = float(spot_fijo_eur.replace('.', '').replace(',', '.'))
    except (ValueError, AttributeError):
        pass

# IPC: Primer valor es el spot_fijo (para calcular IPC en segunda fila)
uf_anterior = spot_fijo_uf_num

for idx, fecha_str in enumerate(fechas_ordenadas):
    row_idx = idx + 3  # Row 3, 4, 5
    data_row = resultados_matriz[fecha_str]
    es_primera_fila = (idx == 0)  # Primera fila tiene spot_fijo
    
    # Columna A: Fecha
    ws.cell(row=row_idx, column=1, value=data_row["fecha"].strftime("%d/%m/%Y"))
    
    # ==== UF (B-E) ====
    # Columna B: Spot UF (solo primera fila)
    if es_primera_fila and spot_fijo_uf_num is not None:
        cell = ws.cell(row=row_idx, column=2, value=spot_fijo_uf_num)
        cell.number_format = '#,##0.0000'
    else:
        ws.cell(row=row_idx, column=2, value="")
    
    # Columna C: Pts Forward UF
    uf_pts_num = None
    if "UF/USD" in data_row:
        uf_pts_str = data_row["UF/USD"]["pts_fwd"]
        try:
            uf_pts_num = float(uf_pts_str.replace(',', '.'))
            cell = ws.cell(row=row_idx, column=3, value=uf_pts_num)
            cell.number_format = '#,##0.0000'
        except (ValueError, AttributeError):
            ws.cell(row=row_idx, column=3, value=uf_pts_str)
    else:
        ws.cell(row=row_idx, column=3, value="")
    
    # Columna D: Precio Forward UF = Spot fijo + Pts
    if spot_fijo_uf_num is not None and uf_pts_num is not None:
        precio_fwd_uf = spot_fijo_uf_num + uf_pts_num
        cell = ws.cell(row=row_idx, column=4, value=precio_fwd_uf)
        cell.number_format = '#,##0.0000'
        
        # Columna E: IPC (solo desde segunda fila en adelante)
        if not es_primera_fila and uf_anterior is not None and precio_fwd_uf != 0:
            ipc = (uf_anterior / precio_fwd_uf) - 1
            cell = ws.cell(row=row_idx, column=5, value=ipc)
            cell.number_format = '0.00%'
        else:
            ws.cell(row=row_idx, column=5, value="")
        
        uf_anterior = precio_fwd_uf  # Actualizar para siguiente fila
    else:
        ws.cell(row=row_idx, column=4, value="")
        ws.cell(row=row_idx, column=5, value="")
    
    # ==== USD (F-H) ====
    # Columna F: Spot USD (solo primera fila)
    if es_primera_fila and spot_fijo_usd_num is not None:
        cell = ws.cell(row=row_idx, column=6, value=spot_fijo_usd_num)
        cell.number_format = '#,##0.00'
    else:
        ws.cell(row=row_idx, column=6, value="")
    
    # Columna G: Pts Forward USD
    usd_pts_num = None
    if "USD/CLP" in data_row:
        pts_str = data_row["USD/CLP"]["pts_fwd"]
        try:
            usd_pts_num = float(pts_str.replace('.', '').replace(',', '.'))
            cell = ws.cell(row=row_idx, column=7, value=usd_pts_num)
            cell.number_format = '#,##0.00'
        except (ValueError, AttributeError):
            ws.cell(row=row_idx, column=7, value=pts_str)
    else:
        ws.cell(row=row_idx, column=7, value="")
    
    # Columna H: Precio Forward USD = Spot fijo + Pts
    if spot_fijo_usd_num is not None and usd_pts_num is not None:
        precio_fwd_usd = spot_fijo_usd_num + usd_pts_num
        cell = ws.cell(row=row_idx, column=8, value=precio_fwd_usd)
        cell.number_format = '#,##0.00'
    else:
        ws.cell(row=row_idx, column=8, value="")
    
    # ==== EUR (I-K) ====
    # Columna I: Spot EUR (solo primera fila)
    if es_primera_fila and spot_fijo_eur_num is not None:
        cell = ws.cell(row=row_idx, column=9, value=spot_fijo_eur_num)
        cell.number_format = '#,##0.00'
    else:
        ws.cell(row=row_idx, column=9, value="")
    
    # Columna J: Pts Forward EUR
    eur_pts_num = None
    if "EUR/CLP" in data_row:
        pts_str = data_row["EUR/CLP"]["pts_fwd"]
        try:
            eur_pts_num = float(pts_str.replace('.', '').replace(',', '.'))
            cell = ws.cell(row=row_idx, column=10, value=eur_pts_num)
            cell.number_format = '#,##0.00'
        except (ValueError, AttributeError):
            ws.cell(row=row_idx, column=10, value=pts_str)
    else:
        ws.cell(row=row_idx, column=10, value="")
    
    # Columna K: Precio Forward EUR = Spot fijo + Pts
    if spot_fijo_eur_num is not None and eur_pts_num is not None:
        precio_fwd_eur = spot_fijo_eur_num + eur_pts_num
        cell = ws.cell(row=row_idx, column=11, value=precio_fwd_eur)
        cell.number_format = '#,##0.00'
    else:
        ws.cell(row=row_idx, column=11, value="")

# ===== FORMATO =====
# Anchos de columna
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 14
ws.column_dimensions['H'].width = 14
ws.column_dimensions['I'].width = 14
ws.column_dimensions['J'].width = 14
ws.column_dimensions['K'].width = 14

# Bordes
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=11):
    for cell in row:
        cell.border = thin_border

# ===== NOTA EXPLICATIVA (2 filas debajo de la tabla) =====
nota_row = ws.max_row + 2  # 2 filas después del final de la tabla

# Combinar celdas A-K para la nota
ws.merge_cells(start_row=nota_row, start_column=1, end_row=nota_row, end_column=11)

# Texto de la nota
nota_cell = ws.cell(row=nota_row, column=1)
nota_cell.value = (
    "Nota: El Precio Forward se calcula como Spot (fila 1) + Puntos Forward (cada fila). "
    "El IPC se calcula como (Precio Forward anterior / Precio Forward actual) - 1."
)

# Estilo: Calibri, cursiva, tamaño 9, alineado a la izquierda
nota_cell.font = Font(name='Calibri', size=9, italic=True, color="666666")
nota_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Sin borde o con borde sutil
nota_cell.border = Border()

# ===== GUARDAR =====
output_path = Path(r"C:\Users\klaus\OneDrive\Documentos")
filename = f"Forward_TEST_SPOTFIJO_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
filepath = output_path / filename

wb.save(filepath)

print(f"\n✅ Excel TEST guardado en: {filepath}")
print(f"\n📊 NUEVA Estructura con spot_fijo:")
print(f"   Row 1: UF y IPC (B-E merged) | USD (F-H merged) | EUR (I-K merged)")
print(f"   Row 2: Fecha | Spot | Pts.Fwd | Precio Fwd | IPC | Spot | Pts.Fwd | Precio Fwd | Spot | Pts.Fwd | Precio Fwd")
print(f"   Row 3 (1era fecha): {fechas_ordenadas[0]} | {spot_fijo_uf} | pts | precio | - | {spot_fijo_usd} | pts | precio | {spot_fijo_eur} | pts | precio")
print(f"   Row 4-5 (2da-3ra): Fechas | - | pts | precio | IPC% | - | pts | precio | - | pts | precio")
print(f"\n🔍 Validar:")
print(f"   ✓ Row 3 (primera fecha) tiene Spots fijos + Pts + Precio Forward")
print(f"   ✓ Row 4-5 (fechas siguientes) tienen solo Pts + Precio Forward (Spot vacío)")
print(f"   ✓ Precio Forward = Spot fijo (Row 3) + Pts Forward (todas las filas)")
print(f"   ✓ IPC calculado desde Row 4 en adelante (entre precios forward consecutivos)")
print(f"   ✓ Nota explicativa agregada 2 filas debajo (en cursiva, Calibri, gris)")
