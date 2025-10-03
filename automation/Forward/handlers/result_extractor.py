"""
ResultExtractor - Responsabilidad: Extracción de resultados de cálculos

Principio de Responsabilidad Única (SRP):
- Extraer valores de inputs de resultados
- Formatear datos extraídos
- Guardar resultados en Excel
"""

from datetime import datetime
from typing import Callable, Dict, List, TYPE_CHECKING
import time

if TYPE_CHECKING:
    from selenium.webdriver.remote.webdriver import WebDriver

from selenium.webdriver.common.by import By


class ResultExtractor:
    """
    Extractor especializado en resultados de cálculos Forward.
    
    Responsabilidades:
    - Extraer Precio Spot Bid
    - Extraer Puntos Forward Bid
    - Guardar resultados en Excel
    """
    
    # XPaths de campos de resultados (ACTUALIZADOS - Columna correcta)
    # Paridades normales (USD/CLP, EUR/CLP) - app-currency-forward
    PRECIO_SPOT_BID_XPATH = "/html/body/app-root/app-home/div[2]/app-calculator/div/div/div/div[2]/app-currency-forward/app-calculator-forward/div[2]/fieldset/div[2]/div[2]/div/label/app-input-decimal/input"
    
    PUNTOS_FORWARD_BID_XPATH = "/html/body/app-root/app-home/div[2]/app-calculator/div/div/div/div[2]/app-currency-forward/app-calculator-forward/div[2]/fieldset/div[3]/div[2]/div/span/app-input-decimal-ptos-fwd/input"
    
    # UF - app-inflation-forward (interfaz diferente)
    UF_PRECIO_SPOT_BID_XPATH = "/html/body/app-root/app-home/div[2]/app-calculator/div/div/div/div[2]/app-inflation-forward/app-calculator-forward/div[2]/fieldset/div[2]/div[2]/div/label/app-input-decimal/input"
    
    UF_PUNTOS_FORWARD_BID_XPATH = "/html/body/app-root/app-home/div[2]/app-calculator/div/div/div/div[2]/app-inflation-forward/app-calculator-forward/div[2]/fieldset/div[3]/div[2]/div/span/app-input-decimal-ptos-fwd/input"
    
    def __init__(self, driver: "WebDriver", log_callback: Callable):
        """
        Args:
            driver: Instancia de Selenium WebDriver
            log_callback: Función para logging
        """
        self.driver = driver
        self.log = log_callback
    
    def extraer_resultados(self, mes_vencimiento: datetime, es_uf: bool = False) -> Dict:
        """
        Extrae los resultados de un cálculo.
        
        Args:
            mes_vencimiento: Mes del cálculo
            es_uf: True si es UF (usa XPaths de app-inflation-forward), False para USD/EUR
            
        Returns:
            Diccionario con estructura:
            {
                "mes": "Enero",
                "año": 2025,
                "mes_numero": 1,
                "precio_spot_bid": "958,35",
                "puntos_forward_bid": "123,45"
            }
        """
        try:
            tipo_paridad = "UF (app-inflation-forward)" if es_uf else "USD/EUR (app-currency-forward)"
            self.log(f"[ResultExtract] 📊 Extrayendo resultados ({tipo_paridad})...", level="INFO")
            
            time.sleep(3)  # Esperar a que se carguen los resultados
            
            # Seleccionar XPaths según tipo de paridad
            xpath_spot = self.UF_PRECIO_SPOT_BID_XPATH if es_uf else self.PRECIO_SPOT_BID_XPATH
            xpath_pts = self.UF_PUNTOS_FORWARD_BID_XPATH if es_uf else self.PUNTOS_FORWARD_BID_XPATH
            
            # Reintentar si falla
            for intento in range(3):
                try:
                    precio_spot_bid = self._extraer_valor(xpath_spot)
                    puntos_forward_bid = self._extraer_valor(xpath_pts)
                    break
                except Exception:
                    if intento < 2:
                        time.sleep(2)
                    else:
                        raise
            
            # Construir resultado
            resultado = {
                "mes": mes_vencimiento.strftime("%B"),
                "año": mes_vencimiento.year,
                "mes_numero": mes_vencimiento.month,
                "precio_spot_bid": precio_spot_bid,
                "puntos_forward_bid": puntos_forward_bid
            }
            
            self.log(f"[ResultExtract]    ├─ Precio Spot Bid: {precio_spot_bid}", level="INFO")
            self.log(f"[ResultExtract]    └─ Puntos Forward Bid: {puntos_forward_bid}", level="INFO")
            self.log(
                f"[ResultExtract] ✅ Resultados extraídos para {mes_vencimiento.strftime('%B %Y')}", 
                level="EXITO"
            )
            
            return resultado
            
        except Exception as e:
            self.log(f"[ResultExtract] ❌ Error extrayendo resultados: {e}", level="ERROR")
            raise
    
    def guardar_resultados_excel(
        self, 
        resultados: List[Dict], 
        directorio_salida: str
    ):
        """
        Guarda los resultados en un archivo Excel.
        
        Args:
            resultados: Lista de diccionarios con resultados
            directorio_salida: Directorio donde guardar el archivo
        """
        try:
            import pandas as pd
            from pathlib import Path
            
            self.log("[ResultExtract] 💾 Guardando resultados en Excel...", level="INFO")
            
            if not resultados:
                self.log("[ResultExtract] ⚠️ No hay resultados para guardar", level="ADVERTENCIA")
                return
            
            # Crear directorio si no existe
            output_path = Path(directorio_salida)
            if not output_path.exists():
                output_path.mkdir(parents=True, exist_ok=True)
            
            # Crear DataFrame
            df = pd.DataFrame(resultados)
            df = df.sort_values(by=['año', 'mes_numero'])
            
            # Generar nombre de archivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"forward_resultados_{timestamp}.xlsx"
            filepath = output_path / filename
            
            # Guardar a Excel
            df.to_excel(
                filepath, 
                index=False, 
                columns=['mes', 'año', 'precio_spot_bid', 'puntos_forward_bid']
            )
            
            self.log(f"[ResultExtract] ✅ Guardado en: {filepath}", level="EXITO")
            self.log(f"[ResultExtract] 📊 Total de filas: {len(df)}", level="INFO")
            
        except Exception as e:
            self.log(f"[ResultExtract] ❌ Error guardando Excel: {e}", level="ERROR")
            raise
    
    def guardar_resultados_excel_matriz(
        self, 
        resultados_matriz: Dict[str, Dict],
        paridades: List[str],
        directorio_salida: str
    ):
        """
        Guarda resultados en formato de matriz con headers agrupados.
        
        Args:
            resultados_matriz: Dict[fecha_str, Dict[paridad, Dict[spot_fijo, pts_fwd]]]
            paridades: Lista de paridades procesadas
            directorio_salida: Directorio donde guardar
            
        NUEVO Formato Excel:
            Row 1: "UF y IPC" (B-E merged) | "USD" (F-H merged) | "EUR" (I-K merged)
            Row 2: Fecha | Spot | Pts.Fwd | Precio Fwd | IPC | Spot | Pts.Fwd | Precio Fwd | Spot | Pts.Fwd | Precio Fwd
            Row 3 (1era fecha): 30/10 | 0.0254 | 0.0001 | 0.0255 | - | 920.50 | 12.30 | 932.80 | 1050.20 | 15.40 | 1065.60
            Row 4+ (siguientes): 28/11 | - | 0.0002 | 0.0256 | 1.2% | - | 13.10 | 933.60 | - | 16.20 | 1066.40
            
        NUEVA Lógica:
            - Spot se extrae SOLO 1 vez (iteración 1) por paridad → spot_fijo
            - Primera fila (Row 3): Muestra Spot fijo + Pts Forward + Precio Forward calculado
            - Filas siguientes (Row 4+): Solo Pts Forward + Precio Forward (Spot vacío, se usa el fijo de Row 3)
            - Precio Forward = spot_fijo + pts_fwd (mismo Spot fijo para todas las filas)
            - IPC calculado desde Row 4 en adelante (entre precios forward consecutivos)
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
            from pathlib import Path
            
            self.log("[ResultExtract] 💾 Creando Excel con NUEVO formato matriz...", level="INFO")
            
            if not resultados_matriz:
                self.log("[ResultExtract] ⚠️ No hay resultados para guardar", level="ADVERTENCIA")
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Forward Results"
            
            # Separar UF de otras paridades
            paridades_normales = [p for p in paridades if p != "UF"]
            tiene_uf = "UF" in paridades
            
            self.log(f"[ResultExtract] 📋 Formato: UF={tiene_uf}, Paridades normales={paridades_normales}", level="INFO")
            
            # ===== ROW 1: HEADERS PRINCIPALES (merged) =====
            # Columna A (Fecha) - vacía en Row 1
            
            # Columnas B-E: "UF y IPC" (4 columnas: Spot UF, Pts Fwd UF, Precio Fwd UF, IPC)
            cell = ws.cell(row=1, column=2, value="UF y IPC")
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=5)
            
            # Columnas F-H: "USD" (3 columnas: Spot, Pts Fwd, Precio Fwd)
            if "USD/CLP" in paridades_normales:
                cell = ws.cell(row=1, column=6, value="USD")
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=8)
            
            # Columnas I-K: "EUR" (3 columnas: Spot, Pts Fwd, Precio Fwd)
            if "EUR/CLP" in paridades_normales:
                cell = ws.cell(row=1, column=9, value="EUR")
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=11)
            
            # ===== ROW 2: SUB-HEADERS =====
            header_style = {
                'font': Font(bold=True, size=10),
                'fill': PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
                'alignment': Alignment(horizontal='center', vertical='center')
            }
            
            # Columna A: Fecha
            cell = ws.cell(row=2, column=1, value="Fecha")
            for k, v in header_style.items(): setattr(cell, k, v)
            
            # Columna B: Spot (UF)
            cell = ws.cell(row=2, column=2, value="Spot")
            for k, v in header_style.items(): setattr(cell, k, v)
            
            # Columna C: Pts. Forward (UF)
            cell = ws.cell(row=2, column=3, value="Pts. Forward")
            for k, v in header_style.items(): setattr(cell, k, v)
            
            # Columna D: Precio Forward (UF)
            cell = ws.cell(row=2, column=4, value="Precio Forward")
            for k, v in header_style.items(): setattr(cell, k, v)
            
            # Columna E: IPC
            cell = ws.cell(row=2, column=5, value="IPC")
            for k, v in header_style.items(): setattr(cell, k, v)
            
            # USD: Columnas F, G, H
            if "USD/CLP" in paridades_normales:
                cell = ws.cell(row=2, column=6, value="Spot")
                for k, v in header_style.items(): setattr(cell, k, v)
                
                cell = ws.cell(row=2, column=7, value="Pts. Forward")
                for k, v in header_style.items(): setattr(cell, k, v)
                
                cell = ws.cell(row=2, column=8, value="Precio Forward")
                for k, v in header_style.items(): setattr(cell, k, v)
            
            # EUR: Columnas I, J, K
            if "EUR/CLP" in paridades_normales:
                cell = ws.cell(row=2, column=9, value="Spot")
                for k, v in header_style.items(): setattr(cell, k, v)
                
                cell = ws.cell(row=2, column=10, value="Pts. Forward")
                for k, v in header_style.items(): setattr(cell, k, v)
                
                cell = ws.cell(row=2, column=11, value="Precio Forward")
                for k, v in header_style.items(): setattr(cell, k, v)
            
            # ===== EXTRAER SPOTS FIJOS (Iteración 1) =====
            # Buscar en la primera fecha que tenga spot_fijo != None para cada paridad
            spot_fijo_uf = None
            spot_fijo_usd = None
            spot_fijo_eur = None
            
            fechas_ordenadas = sorted(
                resultados_matriz.keys(), 
                key=lambda x: resultados_matriz[x]["fecha"]
            )
            
            for fecha_str in fechas_ordenadas:
                data_row = resultados_matriz[fecha_str]
                
                # Extraer spot_fijo de UF
                if tiene_uf and "UF" in data_row:
                    if data_row["UF"].get("spot_fijo") and spot_fijo_uf is None:
                        spot_fijo_uf = data_row["UF"]["spot_fijo"]
                
                # Extraer spot_fijo de USD/CLP
                if "USD/CLP" in paridades_normales and "USD/CLP" in data_row:
                    if data_row["USD/CLP"].get("spot_fijo") and spot_fijo_usd is None:
                        spot_fijo_usd = data_row["USD/CLP"]["spot_fijo"]
                
                # Extraer spot_fijo de EUR/CLP
                if "EUR/CLP" in paridades_normales and "EUR/CLP" in data_row:
                    if data_row["EUR/CLP"].get("spot_fijo") and spot_fijo_eur is None:
                        spot_fijo_eur = data_row["EUR/CLP"]["spot_fijo"]
            
            self.log(f"[ResultExtract] 🎯 Spots fijos: UF={spot_fijo_uf}, USD={spot_fijo_usd}, EUR={spot_fijo_eur}", level="INFO")
            
            # ===== ROW 3+: FILAS DE DATOS =====
            # Convertir spots fijos a números ANTES del loop
            spot_fijo_uf_num = None
            spot_fijo_usd_num = None
            spot_fijo_eur_num = None
            
            if spot_fijo_uf:
                try:
                    # UF usa formato español: "39.485,65" → remover puntos de miles, cambiar coma a punto
                    spot_fijo_uf_num = float(spot_fijo_uf.replace('.', '').replace(',', '.'))
                except (ValueError, AttributeError):
                    pass
            
            if spot_fijo_usd:
                try:
                    spot_fijo_usd_num = float(spot_fijo_usd.replace('.', '').replace(',', '.'))
                except (ValueError, AttributeError):
                    pass
            
            if spot_fijo_eur:
                try:
                    spot_fijo_eur_num = float(spot_fijo_eur.replace('.', '').replace(',', '.'))
                except (ValueError, AttributeError):
                    pass
            
            # IPC: Primer valor es el spot_fijo (para calcular IPC en segunda fila)
            uf_anterior = spot_fijo_uf_num
            
            for idx, fecha_str in enumerate(fechas_ordenadas):
                row_idx = idx + 3  # Fila en Excel (3, 4, 5, ...)
                data_row = resultados_matriz[fecha_str]
                es_primera_fila = (idx == 0)  # Primera fila tiene spot_fijo
                
                # Columna A: Fecha
                ws.cell(row=row_idx, column=1, value=data_row["fecha"].strftime("%d/%m/%Y"))
                
                # ==== UF y IPC (Columnas B-E) ====
                # Columna B: Spot UF (solo primera fila)
                if es_primera_fila and spot_fijo_uf_num is not None:
                    cell = ws.cell(row=row_idx, column=2, value=spot_fijo_uf_num)
                    cell.number_format = '#,##0.00'
                else:
                    ws.cell(row=row_idx, column=2, value="")
                
                # Columna C: Pts Forward UF
                uf_pts_num = None
                if tiene_uf and "UF" in data_row:
                    uf_pts_str = data_row["UF"]["pts_fwd"]
                    try:
                        # UF puede venir con o sin puntos de miles: "122,3432" o "1.234,56"
                        uf_pts_num = float(uf_pts_str.replace('.', '').replace(',', '.'))
                        cell = ws.cell(row=row_idx, column=3, value=uf_pts_num)
                        cell.number_format = '#,##0.00'
                    except (ValueError, AttributeError):
                        ws.cell(row=row_idx, column=3, value=uf_pts_str)
                        uf_pts_num = None
                else:
                    ws.cell(row=row_idx, column=3, value="")
                
                # Columna D: Precio Forward UF = Spot fijo + Pts
                if spot_fijo_uf_num is not None and uf_pts_num is not None:
                    precio_fwd_uf = spot_fijo_uf_num + uf_pts_num
                    cell = ws.cell(row=row_idx, column=4, value=precio_fwd_uf)
                    cell.number_format = '#,##0.00'
                    
                    # Columna E: IPC (solo desde segunda fila en adelante)
                    if not es_primera_fila and uf_anterior is not None and precio_fwd_uf != 0:
                        ipc = (uf_anterior / precio_fwd_uf) - 1
                        cell = ws.cell(row=row_idx, column=5, value=ipc)
                        cell.number_format = '0.00%'
                    else:
                        ws.cell(row=row_idx, column=5, value="")
                    
                    uf_anterior = precio_fwd_uf  # Actualizar para siguiente fila
                else:
                    ws.cell(row=row_idx, column=4, value="")
                    ws.cell(row=row_idx, column=5, value="")
                
                # ==== USD (Columnas F-H) ====
                # Columna F: Spot USD (solo primera fila)
                if es_primera_fila and spot_fijo_usd_num is not None:
                    cell = ws.cell(row=row_idx, column=6, value=spot_fijo_usd_num)
                    cell.number_format = '#,##0.00'
                else:
                    ws.cell(row=row_idx, column=6, value="")
                
                # Columna G: Pts Forward USD
                usd_pts_num = None
                if "USD/CLP" in paridades_normales and "USD/CLP" in data_row:
                    pts_str = data_row["USD/CLP"]["pts_fwd"]
                    try:
                        usd_pts_num = float(pts_str.replace('.', '').replace(',', '.'))
                        cell = ws.cell(row=row_idx, column=7, value=usd_pts_num)
                        cell.number_format = '#,##0.00'
                    except (ValueError, AttributeError):
                        ws.cell(row=row_idx, column=7, value=pts_str)
                        usd_pts_num = None
                else:
                    ws.cell(row=row_idx, column=7, value="")
                
                # Columna H: Precio Forward USD = Spot fijo + Pts
                if spot_fijo_usd_num is not None and usd_pts_num is not None:
                    precio_fwd_usd = spot_fijo_usd_num + usd_pts_num
                    cell = ws.cell(row=row_idx, column=8, value=precio_fwd_usd)
                    cell.number_format = '#,##0.00'
                else:
                    ws.cell(row=row_idx, column=8, value="")
                
                # ==== EUR (Columnas I-K) ====
                # Columna I: Spot EUR (solo primera fila)
                if es_primera_fila and spot_fijo_eur_num is not None:
                    cell = ws.cell(row=row_idx, column=9, value=spot_fijo_eur_num)
                    cell.number_format = '#,##0.00'
                else:
                    ws.cell(row=row_idx, column=9, value="")
                
                # Columna J: Pts Forward EUR
                eur_pts_num = None
                if "EUR/CLP" in paridades_normales and "EUR/CLP" in data_row:
                    pts_str = data_row["EUR/CLP"]["pts_fwd"]
                    try:
                        eur_pts_num = float(pts_str.replace('.', '').replace(',', '.'))
                        cell = ws.cell(row=row_idx, column=10, value=eur_pts_num)
                        cell.number_format = '#,##0.00'
                    except (ValueError, AttributeError):
                        ws.cell(row=row_idx, column=10, value=pts_str)
                        eur_pts_num = None
                else:
                    ws.cell(row=row_idx, column=10, value="")
                
                # Columna K: Precio Forward EUR = Spot fijo + Pts
                if spot_fijo_eur_num is not None and eur_pts_num is not None:
                    precio_fwd_eur = spot_fijo_eur_num + eur_pts_num
                    cell = ws.cell(row=row_idx, column=11, value=precio_fwd_eur)
                    cell.number_format = '#,##0.00'
                else:
                    ws.cell(row=row_idx, column=11, value="")
            
            # ===== FORMATO =====
            # Anchos de columna
            ws.column_dimensions['A'].width = 12  # Fecha
            ws.column_dimensions['B'].width = 12  # Spot UF
            ws.column_dimensions['C'].width = 14  # Pts Fwd UF
            ws.column_dimensions['D'].width = 14  # Precio Fwd UF
            ws.column_dimensions['E'].width = 10  # IPC
            ws.column_dimensions['F'].width = 14  # Spot USD
            ws.column_dimensions['G'].width = 14  # Pts Fwd USD
            ws.column_dimensions['H'].width = 14  # Precio Fwd USD
            ws.column_dimensions['I'].width = 14  # Spot EUR
            ws.column_dimensions['J'].width = 14  # Pts Fwd EUR
            ws.column_dimensions['K'].width = 14  # Precio Fwd EUR
            
            # Bordes
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=11):
                for cell in row:
                    cell.border = thin_border
            
            # ===== NOTA EXPLICATIVA (2 filas debajo de la tabla) =====
            nota_row = ws.max_row + 2  # 2 filas después del final de la tabla
            
            # Combinar celdas A-K para la nota
            ws.merge_cells(start_row=nota_row, start_column=1, end_row=nota_row, end_column=11)
            
            # Texto de la nota
            nota_cell = ws.cell(row=nota_row, column=1)
            nota_cell.value = (
                "Nota: El Precio Forward se calcula como Spot (primera fila) + Puntos Forward (cada fila). "
                "El IPC se calcula como (Precio Forward anterior / Precio Forward actual) - 1."
            )
            
            # Estilo: Calibri, cursiva, tamaño 9, color gris
            nota_cell.font = Font(name='Calibri', size=9, italic=True, color="666666")
            nota_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Sin borde
            nota_cell.border = Border()
            
            self.log(f"[ResultExtract] 📝 Nota explicativa agregada en fila {nota_row}", level="INFO")
            
            # ===== GUARDAR =====
            output_path = Path(directorio_salida)
            if not output_path.exists():
                output_path.mkdir(parents=True, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            filename = f"Forward_{timestamp}.xlsx"
            filepath = output_path / filename
            
            wb.save(filepath)
            
            self.log(f"[ResultExtract] ✅ Excel matriz guardado: {filepath}", level="EXITO")
            self.log(f"[ResultExtract] 📊 Formato: UF y IPC (B-E) | USD (F-H) | EUR (I-K)", level="INFO")
            self.log(f"[ResultExtract] 📊 Total paridades: {len(paridades)} | Fechas: {len(fechas_ordenadas)}", level="INFO")
            if tiene_uf:
                self.log(f"[ResultExtract]    ⚠️ UF: Spot, Pts Fwd, Precio Fwd calculado", level="INFO")
            self.log(f"[ResultExtract]    ├─ Paridades normales: {', '.join(paridades_normales)}", level="INFO")
            self.log(f"[ResultExtract]    └─ Precio Forward = Spot + Pts Forward (calculado)", level="INFO")
            
            return str(filepath)
            
        except Exception as e:
            self.log(f"[ResultExtract] ❌ Error guardando Excel matriz: {e}", level="ERROR")
            raise
    
    def _extraer_valor(self, xpath: str) -> str:
        """
        Extrae el valor de un input.
        
        Args:
            xpath: XPath del input
            
        Returns:
            Valor del input como string
        """
        elemento = self.driver.find_element(By.XPATH, xpath)
        valor = elemento.get_attribute("value") or elemento.text
        return valor
